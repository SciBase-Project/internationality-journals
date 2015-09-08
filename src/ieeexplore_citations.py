# append paper number to links below to get details
# for eg : http://ieeexplore.ieee.org/xpl/articleDetails.jsp?arnumber=6196220
details_link = 'http://ieeexplore.ieee.org/xpl/articleDetails.jsp?arnumber='
authors_link = 'http://ieeexplore.ieee.org/xpl/abstractAuthors.jsp?arnumber='
references_link = 'http://ieeexplore.ieee.org/xpl/abstractReferences.jsp?arnumber='
citations_link = 'http://ieeexplore.ieee.org/xpl/abstractCitations.jsp?arnumber='

import urllib2
from bs4 import BeautifulSoup


def get_id(link):
    import re

    try :
        searchObj = re.search(r'(.*)arnumber=(\d*).*', link)
        if searchObj:
            return searchObj.group(2)
        else:
            return ""
    except :
        return ""


def get_authors(id):
    link = authors_link + str(id)
    # print "[INFO]  Fetching authors from link " + str(link)

    authors = []
    try:
        page = urllib2.urlopen(link).read()
        soup = BeautifulSoup(page, 'html.parser')

        # print soup.prettify()
        metas = soup.find_all("meta", attrs={"name": "citation_author"})

        for meta in metas:
            if meta['name'] == 'citation_author':
                authors.append(meta['content'])
    except:
        pass

    # print "[DEBUG] ", authors
    return authors


def get_references(id):
    link = references_link + str(id)

    paper = {}
    paper['arnumber'] = id
    paper['authors'] = get_authors(id)
    paper['citations'] = []

    print "[DEBUG]  id : " + str(id) + ", authors : ", paper['authors']

    try:
        page = urllib2.urlopen(link).read()
        soup = BeautifulSoup(page, 'html.parser')

        ol = soup.find('ol', attrs={'class': 'docs'})
        lis = ol.findAll('li')

        for li in lis:
            citation = li.find('a')
            citation_id = get_id(str(citation))
            citation_authors = get_authors(citation_id)

            if citation_id is "" :
                continue

            paper['citations'].append({'arnumber' : citation_id, 'authors' : citation_authors})

    except:
        pass
    
    return paper


all_papers = []


def compute_in_parallel(args):
    for arg in args:
        # print "\n[INFO]  Processing " + str(arg)

        id = get_id(arg)
        if id is not "" :
            all_papers.append( get_references(id) )


def citation_network():
    from openpyxl import load_workbook

    wb = load_workbook(filename='../data/Data3.xlsx', read_only=True)
    ws = wb['Sheet1']


    links = []
    for row in ws.iter_rows('G39:G1038'):
        for cell in row:
            links.append(cell.value)


    import threading
    num_per_thread = 80
    divide = lambda lst, sz: [lst[i : i + sz] for i in range(0, len(lst), sz)]
    divided_links = divide(links, num_per_thread)

    print "[INFO]  Number of threads ", (len(links) / num_per_thread) + 1

    threads = []
    for i in range(len(divided_links)):
        t = threading.Thread(target=compute_in_parallel, args=(divided_links[i],))
        threads.append(t)
        t.start()

    for t in threads :
        t.join()


    print "[INFO]  Number of papers " + str(len(all_papers))
    print "[INFO]  Writing papers to file"

    import json
    file = open("../output/ieee_citation_network.json", "w")
    file.write(json.dumps(all_papers, indent=4))
    file.close()

    print "[INFO]  Done Writing papers to file"


citation_network()
